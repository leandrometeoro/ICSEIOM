"""
04_ingest_alpha3_turismo.py  —  α₃ turismo costeiro (dados reais)

Metodologia:
    α₃(muni) = V_tur(muni) × δ_costeiro(muni)

    V_tur = soma de salarios CNAE 55+56+79+93 (SIDRA 6450, var 662, R$ mil)
        - 55: Alojamento
        - 56: Alimentacao
        - 79: Agencias de viagem e operadores turisticos
        - 93: Atividades esportivas e de recreacao
    Referencia: Freitas, Farias & Carvalho (2022), Tabela 4 "Setores e
    atividades CNAE" do setor Servicos da Economia do Mar e Costeira, em
    "Economia Azul" (DGN/Marinha), p.784-785.

    δ_costeiro = fracao do turismo dependente da costa (0 a 1)
        Fonte primaria: ranking "Sol e Praia" do MTur Atividades Turisticas
            δ = (7 - ranking) / 6   (ranking 1 → δ=1,0; ranking 6 → δ=0,17)
        Fallback: Especializacao Turistica (MTur Categorizacao Simplificado)
            normalizada pelo maximo, cruzada com QL da RTA (Cap. 36, Sanguinet
            & Sass 2022, "Economia Azul", DGN/Marinha).

Fontes de dados:
    1. IBGE SIDRA 6450 (CEMPRE) — salarios por CNAE divisao, todos municipios
    2. MTur Categorizacao Simplificado — Especializacao Turistica + Regiao Turistica
    3. MTur Atividades Turisticas — ranking Sol e Praia por municipio
    4. Cap. 36 Economia Azul (Marinha) — QL turismo por RTA (Apendice A)
"""
import sqlite3, argparse, json, time, os
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import HTTPError

try:
    import openpyxl
except ImportError:
    print("ERRO: pip install openpyxl")
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "db" / "icseiom.db"

# ── CNAEs turismo (codigos na classificacao 12762 do SIDRA) ──────────
CNAES_TURISMO = {
    "117544": "55-Alojamento",
    "117545": "56-Alimentacao",
    "117549": "79-Agencias viagem",
    "117556": "93-Esportes recreacao",
}
CNAE_CODES = ",".join(CNAES_TURISMO.keys())

# ── QL por RTA (Apendice A, Cap. 36, Sanguinet & Sass 2022) ─────────
# Nomes normalizados (upper, sem acento) para facilitar match
QL_POR_RTA = {
    "BAIA DE TODOS OS SANTOS": 1.009,
    "CAMINHOS DO SAO FRANCISCO": 0.970,
    "COSTA DAS BALEIAS": 1.895,
    "COSTA DO CACAU": 4.728,
    "COSTA DO DENDE": 8.427,
    "COSTA DO DESCOBRIMENTO": 23.523,
    "COSTA DOS COQUEIROS": 5.494,
    "COSTA DOS CORAIS": 20.952,
    "REGIAO COSTA DOS CORAIS": 20.952,
    "COSTA NAUTICA COROA DO AVIAO": 0.554,
    "FORTALEZA": 0.638,
    "GRANDE MACEIO": 1.859,
    "HISTORIA E MAR": 1.511,
    "HISTORIAE MAR": 1.511,
    "HISTORICA, DOS ARRECIFES E MANGUEZAIS": 1.593,
    "LAGOAS E MARES DO SUL": 1.371,
    "LITORAL EXTREMO OESTE": 11.205,
    "LITORAL LESTE": 3.160,
    "LITORAL OESTE": 2.062,
    "POLO SAO LUIS": 0.654,
    "POLO AMAZONIA MARANHENSE": 0.223,
    "POLO COSTA BRANCA": 1.127,
    "POLO COSTA DAS DUNAS": 2.530,
    "IGR COSTA DAS DUNAS": 2.530,
    "POLO COSTA DO DELTA": 3.138,
    "POLO COSTA DOS COQUEIRAIS": 1.309,
    "POLO DELTA DAS AMERICAS": 0.172,
    "POLO FLORESTA DOS GUARAS": 0.123,
    "POLO LAGOS E CAMPOS FLORIDOS": 0.082,
    "POLO LENCOIS MARANHENSES": 4.919,
    "POLO LENCOIS & DELTA": 4.919,
    "POLO MUNIN": 0.920,
    "POLO MUNIM": 0.920,
    "ROTA SANHAUA": 0.659,
    "TRILHAS DOS POTIGUARAS": 0.474,
    "TRILHAS DOS POTIGUARA": 0.474,
    "REGIAO TURISTICA AMAZONIA ATLANTICA CAETE": 2.602,
    "REGIAO TURISTICA AMAZONIA ATLANTICA GUAMA": 0.459,
    "REGIAO TURISTICA CABO ORANGE": 3.322,
    "REGIAO TURISTICA DAS FLORESTAS DO MARAJO": 0.044,
    "REGIAO TURISTICA DOS CAMPOS DO MARAJO": 2.802,
    "REGIAO TURISTICA MEIO DO MUNDO": 0.419,
    "BAIXADA VERDE": 0.368,
    "CAMINHOS DA MATA": 0.282,
    "COSTA DA MATA ATLANTICA": 1.324,
    "COSTA DO SOL": 3.197,
    "COSTA DOCE RJ": 0.585,
    "COSTA E DA IMIGRACAO": 1.523,
    "COSTA VERDE": 5.383,
    "LAGAMAR": 2.480,
    "LITORAL NORTE DE SAO PAULO": 9.719,
    "LITORAL NORTE DE SP": 9.719,
    "METROPOLITANA ES": 0.418,
    "METROPOLITANA RJ": 1.202,
    "METROPOLITANA": 1.202,
    "SERRA VERDE IMPERIAL": 2.023,
    "VERDE E DAS AGUAS": 0.976,
    "CAMINHO DOS CANYONS": 0.920,
    "CAMINHO DOS PRINCIPES": 0.514,
    "COSTA DOCE RS": 0.908,
    "COSTA VERDE & MAR": 3.535,
    "ENCANTOS DO SUL SC": 2.376,
    "ENCANTOS DO SUL": 2.376,
    "GRANDE FLORIANOPOLIS": 2.106,
    "LITORAL DO PARANA": 1.789,
    "LITORAL NORTE GAUCHO": 2.247,
    "COSTA DOS RECIFES": 0.908,
}
QL_MAX = max(QL_POR_RTA.values())  # 23.523


def _normalizar(s: str) -> str:
    """Remove acentos e upper para matching."""
    if not s:
        return ""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).upper().strip()


# ── 1. Carregar MTur Categorizacao Simplificado (Especializacao Turistica) ──

def carregar_categorizacao(arquivos: list[str]) -> dict:
    """Retorna {code_muni_int: {nome, uf, rt, cat, espec}}."""
    dados = {}
    for fp in arquivos:
        wb = openpyxl.load_workbook(fp, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = 0
        for i, r in enumerate(rows):
            if r[0] and "Munic" in str(r[0]):
                hdr_idx = i
                break
        for r in rows[hdr_idx + 1:]:
            ibge = int(r[2]) if r[2] else None
            if ibge and ibge not in dados:
                dados[ibge] = {
                    "nome": r[0],
                    "uf": r[1],
                    "rt": str(r[3] or ""),
                    "cat": str(r[4] or ""),
                    "espec": float(r[10]) if r[10] else 0.0,
                }
    return dados


# ── 2. Carregar MTur Atividades Turisticas (ranking Sol e Praia) ─────

def carregar_sol_praia(pasta: str) -> dict:
    """Retorna {nome_muni_UF: ranking_int} ex: {'Porto Seguro_BA': 1}."""
    dados = {}
    for fn in sorted(os.listdir(pasta)):
        if not fn.endswith(".xlsx") or fn.startswith("~"):
            continue
        fp = os.path.join(pasta, fn)
        wb = openpyxl.load_workbook(fp, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[4:]:
            if not r[1]:
                continue
            uf = str(r[0] or "").strip()
            mun = str(r[1] or "").strip()
            # Col 84 = Sol e Praia (ha dois anos), col 90 = ano anterior
            sp = None
            for col in (90, 84):
                if len(r) > col and r[col]:
                    try:
                        sp = int(r[col])
                        break
                    except (ValueError, TypeError):
                        pass
            if sp is not None:
                key = f"{_normalizar(mun)}_{uf}"
                dados[key] = sp
    return dados


# ── 3. Baixar SIDRA 6450 ────────────────────────────────────────────

def _parse_sidra(data: list[dict], resultados: dict) -> int:
    """Acumula valores do JSON SIDRA em resultados. Retorna n de valores ok."""
    ok = 0
    for row in data[1:]:
        code = row.get("D1C", "").strip()
        valor_str = row.get("V", "").strip()
        if not code or valor_str in ("-", "X", "..", "..."):
            continue
        try:
            valor = float(valor_str) * 1000  # SIDRA retorna em R$ mil
        except ValueError:
            continue
        resultados[code] = resultados.get(code, 0.0) + valor
        ok += 1
    return ok


def baixar_sidra_6450(municipios: list[str], ano: int) -> dict:
    """Baixa salarios CNAE turismo do SIDRA 6450 para lista de municipios.
    Retorna {code_muni_str: valor_rs_total}."""
    resultados = {}
    # API tem limite de tamanho de URL; 20 munis por batch e seguro
    batch_size = 20
    batches = [municipios[i:i+batch_size] for i in range(0, len(municipios), batch_size)]

    for bi, batch in enumerate(batches):
        munis_str = ",".join(batch)
        url = (
            f"https://apisidra.ibge.gov.br/values/t/6450"
            f"/n6/{munis_str}/v/662/p/{ano}/c12762/{CNAE_CODES}"
            f"?formato=json"
        )
        print(f"  SIDRA batch {bi+1}/{len(batches)} ({len(batch)} munis)...", end=" ", flush=True)
        try:
            req = Request(url)
            with urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            ok = _parse_sidra(data, resultados)
            print(f"ok ({ok} valores)")
        except (HTTPError, Exception):
            # Batch falhou (provavelmente cod IBGE invalido), tentar individualmente
            print("batch falhou, tentando individual...")
            for code in batch:
                url1 = (
                    f"https://apisidra.ibge.gov.br/values/t/6450"
                    f"/n6/{code}/v/662/p/{ano}/c12762/{CNAE_CODES}"
                    f"?formato=json"
                )
                try:
                    req1 = Request(url1)
                    with urlopen(req1, timeout=30) as resp1:
                        d1 = json.loads(resp1.read().decode("utf-8"))
                    _parse_sidra(d1, resultados)
                except Exception:
                    pass  # codigo invalido, ignora
                time.sleep(0.3)

        if bi < len(batches) - 1:
            time.sleep(1)  # rate limit

    return resultados


# ── 4. Calcular delta_costeiro ───────────────────────────────────────

def calcular_delta(code_muni: int, nome: str, uf: str,
                   categorizacao: dict, sol_praia: dict) -> tuple[float, str]:
    """Retorna (delta_costeiro, via) onde via descreve a fonte usada."""

    # Tentar Sol e Praia primeiro
    key = f"{_normalizar(nome)}_{uf}"
    if key in sol_praia:
        ranking = sol_praia[key]
        delta = (7 - ranking) / 6.0
        return (round(delta, 4), f"Sol e Praia rank={ranking}")

    # Fallback: Especializacao Turistica × QL da RTA
    cat = categorizacao.get(code_muni)
    if cat:
        espec = cat["espec"]
        rt_norm = _normalizar(cat["rt"])
        ql = None
        # Tentar match direto
        if rt_norm in QL_POR_RTA:
            ql = QL_POR_RTA[rt_norm]
        else:
            # Tentar match parcial
            for rta_nome, rta_ql in QL_POR_RTA.items():
                if rta_nome in rt_norm or rt_norm in rta_nome:
                    ql = rta_ql
                    break

        if ql is not None and espec > 0:
            # QL > 1 indica especializacao acima da media nacional
            # Usar QL diretamente normalizado: QL=1 → delta=0.3, QL=5 → delta=0.7, QL>10 → ~1.0
            ql_norm = min(1.0, ql / 10.0)
            # Espec como peso: >5 = turistico forte, >15 = muito forte
            espec_peso = min(1.0, espec / 15.0)
            delta = ql_norm * espec_peso
            return (round(max(0.05, min(1.0, delta)), 4), f"Espec={espec:.1f} QL={ql:.2f}")
        elif espec > 0:
            # So espec, sem QL (municipio nao costeiro ou RTA nao mapeada)
            delta = min(1.0, espec / 15.0) * 0.5  # fator conservador sem QL
            return (round(max(0.05, delta), 4), f"Espec={espec:.1f} sem QL")

    # Default conservador para municipio costeiro sem dados MTur
    return (0.10, "default costeiro")


# ── Pipeline principal ───────────────────────────────────────────────

def ingest_real(ano: int = 2021,
                xlsx_cat: list[str] | None = None,
                pasta_ativ: str | None = None):
    """Ingestao completa de alpha3 com dados reais."""

    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = OFF")
    cur = con.cursor()

    # Municipios costeiros: usar municipios_brasil (443) se disponivel,
    # senao fallback para municipios_costeiros (61)
    n_mb = cur.execute(
        "SELECT COUNT(*) FROM municipios_brasil WHERE is_costeiro=1"
    ).fetchone()[0]
    if n_mb > 0:
        costeiros = cur.execute(
            "SELECT code_muni, nome, uf FROM municipios_brasil WHERE is_costeiro=1"
        ).fetchall()
    else:
        costeiros = cur.execute(
            "SELECT code_muni, nome, uf FROM municipios_costeiros"
        ).fetchall()
    print(f"Municipios costeiros no banco: {len(costeiros)}")

    # 1. Carregar categorizacao MTur
    if xlsx_cat:
        print(f"\nCarregando categorizacao MTur ({len(xlsx_cat)} arquivos)...")
        categorizacao = carregar_categorizacao(xlsx_cat)
        print(f"  {len(categorizacao)} municipios carregados")
    else:
        categorizacao = {}

    # 2. Carregar Sol e Praia
    if pasta_ativ:
        print(f"\nCarregando Atividades Turisticas de {pasta_ativ}...")
        sol_praia = carregar_sol_praia(pasta_ativ)
        print(f"  {len(sol_praia)} municipios com Sol e Praia")
    else:
        sol_praia = {}

    # 3. Baixar SIDRA 6450
    codes_str = [r["code_muni"] for r in costeiros]
    print(f"\nBaixando SIDRA 6450 (CNAE {'+'.join(c.split('-')[0] for c in CNAES_TURISMO.values())}) ano {ano}...")
    sidra = baixar_sidra_6450(codes_str, ano)
    print(f"  {len(sidra)} municipios com dados SIDRA")

    # Se SIDRA nao retornou nada para o ano pedido, tentar ano anterior
    if not sidra and ano > 2019:
        print(f"  Sem dados para {ano}, tentando {ano-1}...")
        sidra = baixar_sidra_6450(codes_str, ano - 1)
        if sidra:
            ano = ano - 1
            print(f"  {len(sidra)} municipios com dados SIDRA ({ano})")

    # 4. Calcular e inserir
    print(f"\nCalculando alpha3...")
    cur.execute("DELETE FROM alpha3_turismo")

    n_inseridos = 0
    n_sem_sidra = 0
    total_vtur = 0.0
    total_alpha3 = 0.0

    for r in costeiros:
        code = r["code_muni"]
        nome = r["nome"]
        uf = r["uf"]
        code_int = int(code)

        # V_tur do SIDRA
        vtur = sidra.get(code, 0.0)
        if vtur == 0:
            n_sem_sidra += 1
            continue

        # delta_costeiro
        delta, via = calcular_delta(code_int, nome, uf, categorizacao, sol_praia)

        # alpha3 = V_tur * delta
        alpha3 = vtur * delta
        fonte = f"SIDRA6450 CNAE55+56+79+93 d={delta:.2f} ({via})"

        cur.execute(
            "INSERT OR REPLACE INTO alpha3_turismo (code_muni, ano, vab_aloj_rs, fonte) "
            "VALUES (?, ?, ?, ?)",
            (code, ano, round(alpha3, 2), fonte),
        )
        n_inseridos += 1
        total_vtur += vtur
        total_alpha3 += alpha3

    # Metadados
    cur.execute(
        """INSERT OR REPLACE INTO metadados_atualizacao
           (fonte, ultima_safra, atualizado_em, url, observacoes,
            nome_humano, orgao, script, observacoes_metodologicas)
           VALUES (?,?,datetime('now'),?,?,?,?,?,?)""",
        (
            "alpha3_turismo",
            str(ano),
            "https://apisidra.ibge.gov.br/values/t/6450",
            f"{n_inseridos} munis, V_tur total R$ {total_vtur/1e6:.1f}M, "
            f"alpha3 total R$ {total_alpha3/1e6:.1f}M",
            "α₃ turismo costeiro",
            "IBGE SIDRA + MTur",
            "scripts/04_ingest_alpha3_turismo.py",
            "V_tur = SIDRA 6450 CNAE 55+56+79+93 (Freitas et al. 2022, Tab.4). "
            "delta_costeiro = Sol e Praia MTur (primario) ou Espec×QL (fallback). "
            "QL por RTA: Sanguinet & Sass 2022, Apendice A, Cap.36 Economia Azul.",
        ),
    )
    con.commit()

    # Resumo
    print(f"\n{'='*60}")
    print(f"alpha3 inseridos: {n_inseridos} municipios (ano {ano})")
    print(f"sem dados SIDRA:  {n_sem_sidra} municipios")
    print(f"V_tur total:      R$ {total_vtur/1e6:,.1f} M")
    print(f"alpha3 total:     R$ {total_alpha3/1e6:,.1f} M")

    # Detalhamento
    rows = cur.execute(
        "SELECT m.nome, m.uf, a.vab_aloj_rs, a.fonte "
        "FROM alpha3_turismo a JOIN municipios_costeiros m USING(code_muni) "
        "WHERE a.ano=? ORDER BY a.vab_aloj_rs DESC",
        (ano,),
    ).fetchall()
    print(f"\nTop 20:")
    for i, row in enumerate(rows[:20]):
        print(f"  {row['nome']:30s} {row['uf']}  R$ {row['vab_aloj_rs']:>14,.2f}  {row['fonte'][:50]}")

    con.close()
    return n_inseridos


# ── ingest_demo (mantido para compatibilidade) ──────────────────────

BASE = {"Norte": 210, "Nordeste": 340, "Sudeste": 780, "Sul": 520}
HOT = {
    "3304557": 4.8, "3301009": 5.5, "3300456": 7.2, "2304400": 3.2,
    "2611606": 2.9, "2704302": 3.3, "4209102": 5.0, "2933307": 8.5,
    "2927408": 2.8, "3548500": 3.7,
}

def ingest_demo(ano=2024):
    con = sqlite3.connect(DB)
    cur = con.cursor()
    muns = cur.execute(
        "SELECT code_muni, regiao, pop_2022 FROM municipios_costeiros"
    ).fetchall()
    cur.execute("DELETE FROM alpha3_turismo WHERE ano=?", (ano,))
    for code, regiao, pop in muns:
        pop = pop or 1
        vab = pop * BASE.get(regiao, 400) * HOT.get(code, 1.0)
        cur.execute(
            "INSERT INTO alpha3_turismo (code_muni, ano, vab_aloj_rs, fonte) VALUES (?,?,?,?)",
            (code, ano, round(vab, 2), "demo-modelo"),
        )
    cur.execute(
        """INSERT OR REPLACE INTO metadados_atualizacao
           (fonte, ultima_safra, atualizado_em, url, observacoes)
           VALUES (?,?,datetime('now'),?,?)""",
        ("alpha3_turismo", str(ano),
         "https://apisidra.ibge.gov.br/values/t/6450", "demo"),
    )
    con.commit()
    n = cur.execute(
        "SELECT COUNT(*) FROM alpha3_turismo WHERE ano=?", (ano,)
    ).fetchone()[0]
    con.close()
    print(f"[OK] α₃ demo {ano}: {n} municipios")


# ── CLI ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Ingestao alpha3 turismo costeiro")
    p.add_argument("--ano", type=int, default=2021,
                   help="Ano do SIDRA 6450 (default 2021, ultimo disponivel)")
    p.add_argument("--demo", action="store_true",
                   help="Usar dados sinteticos (demo)")
    p.add_argument("--cat", nargs="+",
                   help="Arquivos XLSX Categorizacao Simplificado MTur")
    p.add_argument("--ativ", type=str,
                   help="Pasta com XLSX Atividades Turisticas MTur")
    args = p.parse_args()

    if args.demo:
        ingest_demo(args.ano)
    else:
        # Defaults para arquivos locais
        cat_default = [
            "/Users/leandro/Downloads/Relatorio de Categorização Simplificado.xlsx",
            "/Users/leandro/Downloads/Relatorio de Categorização Simplificado (2).xlsx",
            "/Users/leandro/Downloads/Relatorio de Categorização Simplificado (3).xlsx",
            "/Users/leandro/Downloads/Relatorio de Categorização Simplificado (4).xlsx",
        ]
        ativ_default = "/tmp/mtur_atividades"

        cat_files = args.cat or cat_default
        ativ_pasta = args.ativ or ativ_default

        # Verificar arquivos
        cat_exist = [f for f in cat_files if os.path.exists(f)]
        if not cat_exist:
            print("AVISO: nenhum arquivo de categorizacao encontrado, sem delta MTur")
        ativ_exist = os.path.isdir(ativ_pasta)
        if not ativ_exist:
            print(f"AVISO: pasta {ativ_pasta} nao encontrada, sem Sol e Praia")

        ingest_real(
            ano=args.ano,
            xlsx_cat=cat_exist or None,
            pasta_ativ=ativ_pasta if ativ_exist else None,
        )
